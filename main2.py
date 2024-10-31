from config import DNA_FQDN, DNA_PORT, DNA_USER, DNA_PASS, DNA_AUTH_API, DNA_DEVICE_API, DNA_INTERFACE_API
import requests
import logging
from requests.auth import HTTPBasicAuth
import urllib3
from datetime import datetime
from prettytable import PrettyTable
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from sys import exit
import os

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # Silence SSL Certificate warnings

current_datetime = datetime.now()  # Get the current datetime
todays_date = current_datetime.strftime("%d/%m/%Y")  # Format as dd/mm/yyyy

header_row = ['Date', 'Switch Name', 'IP Address', 'UP Access Ports', 'UP Module Ports', 'Total UP Ports', 'Total DOWN Ports', 'Admin DOWN Ports', 'Total Ports']

device_table = PrettyTable(header_row)
device_table.padding_width = 1

headers = {'content-type': "application/json", 'x-auth-token': ""}

def dnac_login():
    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_AUTH_API}"
    try:
        response = requests.request("POST", url, auth=HTTPBasicAuth(DNA_USER, DNA_PASS),
                                         headers=headers, verify=False)
    except requests.exceptions.ConnectionError:
        print("Unable to connect to address ", url)
        exit(1)

    if response.status_code != 200:
        print("Login failed. Status code {}".format(response.status_code))

    try:
        token = response.json()["Token"]
        print("Your token is {}".format(token))
        return token
    except KeyError:
        print("No token found in authentication response.")
        print("Response body: ")
        print(response.text)
        exit(1)

def network_switches(token):
    ids = []
    hostnames = []
    platforms = []
    ip_addresses = []
    switch_details = []

    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_DEVICE_API}"
    headers["x-auth-token"] = token

    try:
        response = requests.get(url, headers=headers, verify=False)
        output = response.json().get('response', [])
    except Exception as e:
        logging.error(f"Error fetching device details: {e}")
        return []  # Return an empty list if there's an error

    for device in output:
        if "switch" in device.get("family", "").lower() or "switch" in device.get("type", "").lower():
            ids.append(device.get("id"))
            hostnames.append(device.get("hostname"))
            platforms.append(device.get("platformId"))
            ip_addresses.append(device.get("managementIpAddress"))

    for id, host, platform, ip in zip(ids, hostnames, platforms, ip_addresses):
        switch_details.append((id, host, platform, ip))

    return switch_details

def network_interfaces(token, hostname, id, series):
    total_up = []
    total_down = []
    total_admin_down = []  # Tracking adminStatus "DOWN" interfaces
    total_ports = []
    switch_info = []
    module_ports = []
    access_ports = []

    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_INTERFACE_API}{id}"
    headers["x-auth-token"] = token

    try:
        response = requests.get(url, headers=headers, verify=False)
        output = response.json().get('response', [])
    except Exception as e:
        logging.error(f"Error fetching interfaces for device ID {id}: {e}")
        return 0, 0, 0, [], 0, 0

    for interface in output:
        if isinstance(interface, dict) and interface.get("interfaceType") == "Physical":
            if "GigabitEthernet0/0" != interface.get("portName") and "Bluetooth" not in interface.get("portName"):
                total_ports.append((interface.get("portName"), interface.get("status")))
                if interface.get("adminStatus") == "UP" and interface.get("status") == "up":
                    total_up.append((interface.get("portName"), interface.get("status")))
                elif interface.get("adminStatus") == "DOWN":
                    total_admin_down.append((interface.get("portName"), interface.get("adminStatus")))

    for interface in output:
        if interface.get("status") == "down" and interface.get("interfaceType") == "Physical":
            if "GigabitEthernet0/0" != interface.get("portName") and "Bluetooth" not in interface.get("portName"):
                total_down.append((interface.get("portName"), interface.get("status")))

    switch_info.append(hostname)  # Keep hostname for clarity (used later)
    switch_info.append(series)  # Keep the series (model) for the table

    return len(total_up), len(total_down), len(total_ports), switch_info, len(total_admin_down), len(module_ports), len(access_ports)

if __name__ == "__main__":
    print("Getting DNA Auth Token ...")
    login = dnac_login()

    print("Searching DNA Center Inventory for the following switches...")
    switches = network_switches(login)

    print("Updating or creating Excel file...")
    table_data = []

    print("Generating detailed report for each switch...")
    for id, hostname, series, ip_address in switches:
        interfaces = network_interfaces(login, hostname, id, series)
        up, down, total, info, admin_down, modules, access = interfaces

        # Add current datetime to the row
        row = [todays_date, hostname, ip_address, access, modules, up, down, admin_down, total]
        table_data.append(row)

    # Check if the Excel file exists
    excel_file = 'port-report.xlsx'
    if os.path.exists(excel_file):
        # Load the existing workbook
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    else:
        # Create a new workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Switch Report"
        sheet.append(header_row)  # Add header row only if file doesn't exist
        
    # Get the number of rows before adding new data
    initial_max_row = sheet.max_row

    # Add new data to the sheet
    for row in table_data:
        sheet.append(row)

    # Apply date style to the newly added rows (ensure correct index for new data)
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
    for row_idx in range(initial_max_row + 1, sheet.max_row + 1):
        sheet[f"A{row_idx}"].style = date_style

    # Save the Excel file
    workbook.save(excel_file)

    print(f"Excel report saved: {excel_file}")
