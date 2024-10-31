from config import  DNA_FQDN, DNA_PORT, DNA_USER, DNA_PASS, DNA_AUTH_API, DNA_DEVICE_API, DNA_INTERFACE_API
import requests
import json
import logging
from requests.auth import HTTPBasicAuth
import urllib3
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
from sys import exit
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # Silence the insecure warning debido a SSL Certificate

today = date.today()  # Set date and time
todays_date = today.strftime("%d/%m/%Y")  # Set date format to dd/mm/yyyy

headers = {'content-type': "application/json", 'x-auth-token': ""}

# Function to authenticate with DNA Center and get token
def dnac_login():
    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_AUTH_API}"
    try:
        response = requests.request("POST", url, auth=HTTPBasicAuth(DNA_USER, DNA_PASS),
                                    headers=headers, verify=False)
    except requests.exceptions.ConnectionError:
        print("Unable to connect to address ", url)
        exit(1)

    if response.status_code != 200:
        print("Login failed. Status code {}".format(response.status_code))
        exit(1)

    try:
        token = response.json()["Token"]
        return token
    except KeyError:
        print("No token found in authentication response.")
        print("Response body: ")
        print(response.text)
        exit(1)

# Function to retrieve switch details
def network_switches(token):
    switch_details = []
    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_DEVICE_API}"
    headers["x-auth-token"] = token

    try:
        response = requests.get(url, headers=headers, verify=False)
        devices = response.json().get('response', [])
    except Exception as e:
        logging.error(f"Error fetching device details: {e}")
        return []

    # Print total number of devices fetched
    print(f"Number of devices fetched: {len(devices)}")

    for device in devices:
        if "switch" in device.get("family", "").lower():
            id = device.get("id")
            hostname = device.get("hostname")
            platform = device.get("platformId")
            ip_address = device.get("managementIpAddress")
            switch_details.append((id, hostname, platform, ip_address))

            # Print each switch's details as they're added
            print(f"Switch fetched: Hostname={hostname}, IP={ip_address}, Platform={platform}")

    return switch_details

# Function to fetch interfaces from a switch
def network_interfaces(token, hostname, id):
    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_INTERFACE_API}{id}"
    headers["x-auth-token"] = token

    try:
        response = requests.get(url, headers=headers, verify=False)
        print(f"Raw response for {hostname}: {response.json()}")  # Print full response for debugging
        interfaces = response.json().get('response', [])
    except Exception as e:
        logging.error(f"Error fetching interfaces for device ID {id}: {e}")
        return []  # Ensure an empty list is returned on error

    not_connected_ports = []
    
    for interface in interfaces:
        if interface.get("status") == "down" and interface.get("interfaceType") == "Physical":
            not_connected_ports.append((interface.get("portName"), interface.get("status")))

            # Print each disconnected port found
            print(f"Port {interface.get('portName')} is down on {hostname}")

    # Print the number of not connected ports found
    print(f"Total not connected ports for {hostname}: {len(not_connected_ports)}")
    
    return not_connected_ports


# Function to save data to an Excel file
def save_to_excel(file_name, header_row, data_rows):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(header_row)  # Write header only once

    for row in data_rows:
        sheet.append(row)

    # Auto-adjust column width
    for col in range(1, len(header_row) + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 15

    # Save the Excel file
    workbook.save(file_name)
    print(f"Excel report updated: {file_name}")

# Main script logic
if __name__ == "__main__":
    print("Getting DNA Auth Token ...")
    token = dnac_login()

    print("Fetching switches...")
    switches = network_switches(token)

    xlsx_file = 'portstoshut.xlsx'

    header_row = ['Date', 'Switch Name', 'IP Address', 'Port', 'Status']
    data_rows = []

    print("Fetching interfaces...")
    for id, hostname, platform, ip_address in switches:
        not_connected_ports = network_interfaces(token, hostname, id)
        for port_name, status in not_connected_ports:
            data_rows.append([todays_date, hostname, ip_address, port_name, status])

    save_to_excel(xlsx_file, header_row, data_rows)
