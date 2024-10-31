from config import DNA_FQDN, DNA_PORT, DNA_USER, DNA_PASS, DNA_AUTH_API, DNA_DEVICE_API, DNA_INTERFACE_API
import requests
from requests.auth import HTTPBasicAuth
import urllib3
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from sys import exit
import re

# Disable warnings for insecure HTTPS requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Headers for API requests
headers = {'content-type': "application/json", 'x-auth-token': ""}

# Function to authenticate with DNA Center and get token
def dnac_login():
    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_AUTH_API}"
    try:
        response = requests.post(url, auth=HTTPBasicAuth(DNA_USER, DNA_PASS), headers=headers, verify=False)
    except requests.exceptions.ConnectionError:
        print("Unable to connect to address ", url)
        exit(1)

    if response.status_code != 200:
        print("Login failed. Status code {}".format(response.status_code))
        exit(1)

    try:
        token = response.json()["Token"]
        return token
    except KeyError:
        print("No token found in authentication response.")
        print("Response body: ")
        print(response.text)
        exit(1)

# Excel file path
excel_file = "device_uptime_log.xlsx"

# Function to create Excel file with headers if it doesn't exist
def create_excel_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Uptime Log"
    ws.append(["Timestamp", "Device ID", "Uptime (seconds)", "Reset Detected"])
    
    # Set date format
    date_style = NamedStyle(name="datetime", number_format="DD/MM/YYYY HH:MM:SS")
    for cell in ws["A"]:
        cell.style = date_style
    
    wb.save(excel_file)

# Function to log uptime in Excel format
def log_uptime(device_id, uptime, reset_detected=False):
    # Load or create the workbook
    try:
        wb = load_workbook(excel_file)
    except FileNotFoundError:
        create_excel_file()
        wb = load_workbook(excel_file)
    
    ws = wb.active
    
    # Format the timestamp
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws.append([timestamp, device_id, uptime, reset_detected])
    
    wb.save(excel_file)

# Function to convert uptime string to total seconds
def parse_uptime(uptime_str):
    days = hours = minutes = seconds = 0

    # Regex to match the uptime format
    # Match pattern: "57 days, 22:47:25.00"
    match = re.match(r'(?:(\d+) days?, )?(?:(\d+):)?(?:(\d+):)?(\d+)(?:\.\d+)?', uptime_str)

    if match:
        if match.group(1):
            days = int(match.group(1))
        if match.group(2):
            hours = int(match.group(2))
        if match.group(3):
            minutes = int(match.group(3))
        if match.group(4):
            seconds = int(match.group(4))

    # Calculate total seconds
    total_seconds = (days * 86400) + (hours * 3600) + (minutes * 60) + seconds
    return total_seconds


# Function to get all devices and return their IDs
def get_all_devices():
    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_DEVICE_API}"
    headers['x-auth-token'] = dnac_login()  # Update the headers with the token
    response = requests.get(url, headers=headers, verify=False)
    
    if response.status_code == 200:
        devices = response.json().get("response", [])
        device_ids = [device["id"] for device in devices]
        
        # Print all devices found
        for device in devices:
            print(f"Device ID: {device['id']}, Hostname: {device['hostname']}")
        
        return device_ids
    else:
        print(f"Failed to retrieve devices. Status code: {response.status_code}")
        return []

# Function to get device details including uptime
def get_device_detail(device_id):
    url = f"https://{DNA_FQDN}:{DNA_PORT}{DNA_DEVICE_API}/{device_id}"
    response = requests.get(url, headers=headers, verify=False)
    response.raise_for_status()
    return response.json()

# Function to check for reset based on uptime drops
def check_for_reset(device_id, current_uptime):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Check the last row for previous uptime of the device
        previous_uptime = None
        for row in reversed(list(ws.iter_rows(values_only=True))):
            if row[1] == device_id:
                previous_uptime = int(row[2])
                break
    except FileNotFoundError:
        previous_uptime = None

    # Detect reset if current uptime is lower than previous uptime
    if previous_uptime is not None and current_uptime < previous_uptime:
        log_uptime(device_id, current_uptime, reset_detected=True)
    else:
        log_uptime(device_id, current_uptime, reset_detected=False)

# Main monitoring function
def monitor_devices():
    # Get the list of device IDs from the API
    device_list = get_all_devices()

    for device_id in device_list:
        device_data = get_device_detail(device_id).get("response", {})
        if device_data:
            # Convert uptime string to total seconds
            current_uptime = parse_uptime(device_data.get("upTime", "0 days, 0:0:0.00"))
            # Check for reset and log uptime
            check_for_reset(device_id, current_uptime)

# Run the monitoring function
monitor_devices()
